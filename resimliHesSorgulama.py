
import subprocess
from threading import Thread
from PyQt5.QtGui import QColor, QPixmap
from PyQt5.QtWidgets import QFileDialog, QHeaderView, QLabel, QMessageBox, QMainWindow, QApplication, QDialog, QProgressDialog, QPushButton, QTableWidgetItem
from PyQt5 import uic
from PyQt5.QtCore import QTimer, Qt
import sys, time, os
from selenium.webdriver import Firefox
from selenium.webdriver.firefox.options import Options
from openpyxl import load_workbook
import xlsxwriter
import veritabani as vt


PHOTO_PLACEHOLDER = "resim10.jpg"
TIMEOUT = 15

TC_NO_INDEX = 0
AD_INDEX = 1
SOYAD_INDEX = 2
OKUL_NO_INDEX = 3
SINIF_INDEX = 4
RESIM_INDEX = 5
HES_KODU_INDEX = 6
DURUM_INDEX = 7



class MWindow(QMainWindow):
    
    
    def __init__(self):
        super(MWindow, self).__init__()

        uic.loadUi("ui/anasayfa.ui", self) 
        self.setFixedSize(700,600)

        self.labelHata.setText("")

        self.btnFormuTemizle.clicked.connect(self.formuTemizle)
        self.btnSorgula.clicked.connect(self.hesSorgula)
        
        self.btnKayitEkle.clicked.connect(lambda : self.ekleGuncelle("ekle"))
        self.btnGuncelle.clicked.connect(self.btnGuncelleClicked)
    
        self.btnSil.clicked.connect (self.sil)
        self.btnHepsiniSil.clicked.connect(self.veritabaniniTemizle)
        #self.tcNoListWidget.itemClicked.connect(self.tcNoItemClicked)
        
        self.sinifIslemleri()
        self.btnSinifSorgula.clicked.connect(self.btnSinifSorgulaClicked)
    
        self.dosyaSecButton.clicked.connect(self.dosyaSec)
        self.dizinSecButton.clicked.connect(self.dizinSec)
        self.topluKaydetButton.clicked.connect(self.topluKaydet)
        
        self.btnSinifSil.clicked.connect(self.sinifSilClicked)
        
        self.resim.setPixmap(QPixmap(PHOTO_PLACEHOLDER))
        self.show()
        
        #Bu gizli butonlar thread içinden arayüzde yapılacak değişiklikler için eklendi.
        #Doğru yol olmayabilir ama işe yarıyor :)
        self.animasyonKapatButon = QPushButton(self)
        self.animasyonKapatButon.clicked.connect(self.yukleniyorAnimasyonuKapat)
        self.animasyonKapatButon.hide()
        self.tabloGonderButton = QPushButton(self)
        self.tabloGonderButton.clicked.connect(self.tabloIsle)
        self.tabloGonderButton.hide()
        self.hataMesajiButton = QPushButton(self)
        self.hataMesajiButton.clicked.connect(self.hataMesaji)
        self.hataMesajiButton.hide()
        self.btnSinifIslemleri= QPushButton(self)
        self.btnSinifIslemleri.clicked.connect(self.sinifIslemleri)
        self.btnSinifIslemleri.hide()
        
        self.mesaj = ""
        self.timer=QTimer(self)


    def hataMesaji(self):
        self.timer.stop()
        mesaj = self.mesaj
        self.labelHata.setText(mesaj)
        
        self.timer.timeout.connect(self.hataMesajiKapat)
        self.timer.start(5000)
        
    def hataMesajiKapat(self):
        self.labelHata.setText("")
        self.timer.stop()
    
    def veritabaniniTemizle(self):
        
        qm = QMessageBox(self)
        mesaj = "Tüm bilgiler silinecek. Onaylıyor musunuz?"
        ret = qm.warning(self,'Hepsini Sil', mesaj, qm.Yes | qm.No)

        if ret == qm.Yes:
            vt.DB.veritabanisil()
            
              
                  
    
    def yukleniyorAnimasyonuBaslat(self, parent = None):
        self.dialogMessage = "İşlem devam ediyor..."
        
        if parent == None:
            self.pb = QProgressDialog(self)
        else:
            self.pb = QProgressDialog(parent)

        pb = self.pb
        pb.setLabelText(self.dialogMessage)
        pb.setStyleSheet("color:blue;background-color:orange;")
        pb.setWindowFlag(Qt.FramelessWindowHint)
        pb.setCancelButton(None)
        pb.setRange(0,0)
        pb.setMinimumDuration(0)
        pb.show()

 
    def yukleniyorAnimasyonuKapat(self):
        self.pb.cancel()
        
    
    def ogrenciGoster(self, ogrenci):
        if len(ogrenci) == 0:
            return

        self.tcNo.setText( ogrenci[TC_NO_INDEX] )
        self.ad.setText( ogrenci[AD_INDEX] )
        self.soyad.setText( ogrenci[SOYAD_INDEX] )
        self.okulNo.setText(str(ogrenci[OKUL_NO_INDEX]))
        self.sinif.setText( ogrenci[SINIF_INDEX] )
        self.resim.setPixmap(ogrenci[RESIM_INDEX])
        self.hesKodu.setText(ogrenci[HES_KODU_INDEX])
        self.durum.setText(str(ogrenci[DURUM_INDEX]))
        
     
        
    def hesSorgula(self):
        self.labelHata.setText("")
 
        #*Tc noya göre veritabanı sorgulamasında birden fazla sonuç gelebilir.
        self.tekHesSorguSonucu = []
        hes_kodu = ""
        hesKoduKayitli = True
        
        if self.okulNo.text() != "":
            no = self.okulNo.text()
            self.formuTemizle()
            
            if not no.isnumeric():
                self.mesaj = "Geçerli okul numarası girmelisiniz!"
                self.hataMesajiButton.clik()
                return
            
            ogrenci = vt.DB.okul_no_ile( int(no) )
            if len(ogrenci) == 0:
                self.mesaj = "Bu numara kayıtlı değil."
                self.hataMesajiButton.click()
                return
            
            self.ogrenciGoster(ogrenci)
            hes_kodu = ogrenci[HES_KODU_INDEX]
        
        
        
        elif self.tcNo.text() != "":       
            ogrenci = vt.DB.tc_no_ile( self.tcNo.text() )
            self.formuTemizle()
            
            if len(ogrenci) == 0:
                self.mesaj = "Bu TC no kayıtlı değil."
                self.hataMesajiButton.click()
                return
            
            self.ogrenciGoster(ogrenci)
            hes_kodu = ogrenci[HES_KODU_INDEX]

        
        elif self.hesKodu.text() != "":
            ogrenci = vt.DB.hes_kodu_ile( self.hesKodu.text() )
            
            if len(ogrenci) == 0:
                hes_kodu = self.hesKodu.text()
                hesKoduKayitli = False
            else:       
                self.formuTemizle()
                self.ogrenciGoster(ogrenci)
                hes_kodu = ogrenci[HES_KODU_INDEX]
        
        
        else:
            self.mesaj = "Okul no, TC no veya HES kodu girmelisiniz!"
            self.hataMesajiButton.click()
            return

        if self.kullaniciAdi.text() != "" and self.sifre != "":
            #Girilen bilgiye göre öğrenci kayıtlıysa gösterildi. Hes kodu sorgulanacak.
            self.yukleniyorAnimasyonuBaslat()    
            
            #Firefox yeni bir threadde açılmalı yoksa arayüzü kilitliyor.
            self.mThread = Thread(target=self.eDevlettenSorgula, args=( hes_kodu , hesKoduKayitli))                                                              
            self.mThread.start()
            
        else:
            #*Kullanıcı adı veya şifresi boş
            self.mesaj = "e-Devlet kullanıcı adı ve şifresini girmelisiniz."
            self.hataMesajiButton.click()
            return

    
    
    def formuTemizle(self):
        self.okulNo.setText("")
        self.ad.setText("")
        self.soyad.setText("")
        self.sinif.setText("")
        self.tcNo.setText("")
        self.hesKodu.setText("")
        self.durum.setText("")
        self.resim.setPixmap(QPixmap(PHOTO_PLACEHOLDER))
    
    def yardim(self):
        yardim = QDialog(self)
        uic.loadUi("ui/ekleGuncelle.ui", self.ekleGuncellePencere)
        
        self.ekleGuncellePencere.setFixedSize(450, 280)
        
    def btnGuncelleClicked(self):
        if self.kullaniciAdi.text() == "" or self.sifre.text() == "":
            self.mesaj = "E-Devlet bilgilerini girmeden güncelleme yapılamaz!"
            self.hataMesajiButton.click()
            return
        
        
        tmp_no = self.okulNo.text()
        ogrenci = []
        if tmp_no != "" and tmp_no.isnumeric():
            ogrenci = vt.DB.okul_no_ile(tmp_no)
            
            
        if len(ogrenci) == 0: 
            
            self.mesaj = "Bu numaraya sahip öğrenci yok!"
            self.hataMesajiButton.click()
        else:
            self.ogrenciGoster(ogrenci)
            self.ekleGuncelle("güncelle")
        
    
    def ekleGuncelle(self, eylem):
        if self.kullaniciAdi.text() == "" or self.sifre.text() == "":
            self.mesaj = "E-Devlet bilgilerini girmeden güncelleme yapılamaz!"
            self.hataMesajiButton.click()
            return
        
        self.eklenenResim = ""
        
        self.ekleGuncellePencere = QDialog(self)
        uic.loadUi("ui/ekleGuncelle.ui", self.ekleGuncellePencere)
        
        self.ekleGuncellePencere.setFixedSize(450, 280)
        self.ekleGuncellePencere.hata.hide() 
         
        
        if eylem=="ekle":
            self.ekleGuncellePencere.setWindowTitle("Ekle")
            self.ekleGuncellePencere.widgetKaydet.setText("Kaydet")
        
        elif eylem=="güncelle":
            self.ekleGuncellePencere.setWindowTitle("Güncelle")
            self.ekleGuncellePencere.widgetKaydet.setText("Güncelle")
            
            self.ekleGuncellePencere.widgetOkulNo.setText(self.okulNo.text())
            self.ekleGuncellePencere.widgetOkulNo.setEnabled(False)
            
            self.ekleGuncellePencere.widgetAdi.setText(self.ad.text())
            self.ekleGuncellePencere.widgetAdi.setEnabled(False)
            
            self.ekleGuncellePencere.widgetSoyadi.setText(self.soyad.text())
            self.ekleGuncellePencere.widgetSoyadi.setEnabled(False)
            
            self.ekleGuncellePencere.widgetSinifi.setText(self.sinif.text())
            self.ekleGuncellePencere.widgetTcKimlikNo.setText(self.tcNo.text())
            self.ekleGuncellePencere.widgetTcKimlikNo.setEnabled(False)
            
            self.ekleGuncellePencere.widgetHesKodu.setText(self.hesKodu.text())
            
            res = self.resim.pixmap()
            if res:
                self.ekleGuncellePencere.widgetResim.setPixmap(res)
            
        
        
        self.ekleGuncellePencere.widgetIptal.clicked.connect(self.ekleGuncellePencere.destroy)
        self.ekleGuncellePencere.widgetKaydet.clicked.connect(self.kaydet)
        self.ekleGuncellePencere.widgetResimEkle.clicked.connect(self.resimEkle)

        
        self.ekleGuncellePencere.exec_()
        
    
    
    def resimEkle(self):
        
        resimAdi, _ = QFileDialog.getOpenFileName(self.ekleGuncellePencere, "Resim seç...", "C:\\")
        try:
            if resimAdi != "":
                mResim = QPixmap(resimAdi)
                self.ekleGuncellePencere.widgetResim.setPixmap(mResim)
                self.eklenenResim = resimAdi

        except:
            return
    
    
    
    def kaydet(self):
        self.ekleGuncellePencere.hata.hide()
        
        kimlikNo = self.ekleGuncellePencere.widgetTcKimlikNo.text()
        hesKodu = self.ekleGuncellePencere.widgetHesKodu.text()
        ad = self.ekleGuncellePencere.widgetAdi.text()
        soyad = self.ekleGuncellePencere.widgetSoyadi.text()
        sinif = self.ekleGuncellePencere.widgetSinifi.text()
        okulNo = self.ekleGuncellePencere.widgetOkulNo.text()
        
        ogrenci = [kimlikNo, ad, soyad,okulNo,sinif, self.eklenenResim, hesKodu, "1"]
        
        if self.ekleGuncellePencere.widgetKaydet.text() == "Kaydet":
            #*Bilgilerde eksiklik varsa durdurulacak.
            bilgilerTam = hesKodu != "" and kimlikNo !="" and ad != "" and self.eklenenResim != "" and soyad != "" and sinif != "" and okulNo != ""
            
            if not bilgilerTam:
                self.ekleGuncellePencere.hata.show()
                self.ekleGuncellePencere.hata.setText("Bilgiler eksik!")
                return

            #Öğrenci zaten kayıtlı mı kontrol ediliyor.
            numaraKayitli = vt.DB.okul_no_ile(ogrenci[OKUL_NO_INDEX])
            tcNoKayitli = vt.DB.tc_no_ile(ogrenci[TC_NO_INDEX])
            if len(numaraKayitli) > 0 or len(tcNoKayitli) > 0:
                self.ekleGuncellePencere.hata.show()
                self.ekleGuncellePencere.hata.setText("Öğrenci zaten kayıtlı. Düzenlemek için Güncelle bölümünü kullanın.")
                return
            

            try:
                th = Thread(target=self.eDevlettenGrupGuncelleSayfasiAc, args=("öğrenciler", True, ogrenci))
                th.start()
            except:
                self.ekleGuncellePencere.hata.show()
                self.ekleGuncellePencere.hata.setText("Öğrenci zaten kayıtlı. Düzenlemek için Güncelle bölümünü kullanın.")    
        
        #Güncelleme sayfasındayız.              
        else:
            bilgilerTam = hesKodu != "" and sinif != ""
            
            if not bilgilerTam:
                self.ekleGuncellePencere.hata.show()
                self.ekleGuncellePencere.hata.setText("Bilgiler eksik!")
                return
            
            #Hes kodu değişmedi, sadece veritabanına yazılacak.
            if self.hesKodu.text() == hesKodu:
                
                if self.eklenenResim != "":
                    vt.DB.toplu_kayit_ekle([ogrenci])
                else:
                    vt.DB.veriguncelle(okulNo, sinif, hesKodu)
            
            #Hes değişti. E-devletten hes kodu eklenecek. Eklenirse veritabanına kaydedilecek.
            else:
                if self.eklenenResim != "":                     
                    th = Thread(target= self.eDevlettenGrupGuncelleSayfasiAc, args=("öğrenciler", True, ogrenci))
                    th.start()
                #Yeni resim eklenmediyse eski öğrenci güncellenecek.
                else:
                    th = Thread(target= self.eDevlettenGrupGuncelleSayfasiAc, args=("öğrenciler", True, ogrenci, True))
                    th.start()
            


        self.ekleGuncellePencere.close()
    
    


    def grubaHesEkle(self, hes, browserAc=True):
        if browserAc:
            if not self.eDevletGirisiYap():
                self.mesaj = "E-Devlet girişi yapılamadı!"
                self.hataMesajiButton.click()
                return

        br = self.br
        
        self.elemanYuklenmesiniBekle(id="hesKodu")
        
        br.find_element_by_id("hesKodu").clear()
        br.find_element_by_id("hesKodu").send_keys(hes)
        br.find_element_by_name("btn").click()
        self.elemanYuklenmesiniBekle(sinif="actionButton")
        
        if len( br.find_elements_by_class_name("confirmContainer") ) > 0:
            return True
        else:
            return False
        
        

    #Öğrenci silmek için
    def sil(self):
        
        okulNo = self.okulNo.text()
        if okulNo != "" and okulNo.isnumeric() :
            
            ogrenci = vt.DB.okul_no_ile(int(okulNo))
            if len(ogrenci) > 0:
                self.ogrenciGoster(ogrenci)

                qm = QMessageBox(self)
                ret = qm.question(self,'Sil', "Öğrenci silisin mi?", qm.Yes | qm.No)
                if ret == qm.Yes:
                    vt.DB.kayitsil(int(okulNo))
                    self.mesaj = "Öğrenci silindi."
                    self.hataMesajiButton.click()

        
            else:
                self.mesaj = "Kayıt yok!"
                self.hataMesajiButton.click()
        
        else:
            self.mesaj = "Sileceğiniz öğrencinin numarasını girmediniz!"
            self.hataMesajiButton.click()       
        
        
        
    def sonucGeldimi(self):
        
        try: 
            br=self.br
            tc_sonu = br.find_elements_by_class_name("compact")[0].find_elements_by_tag_name("dd")[1].get_attribute("innerText")
            ad_soyad = br.find_elements_by_class_name("compact")[0].find_elements_by_tag_name("dd")[0].get_attribute("innerText")
            durum = br.find_elements_by_class_name("compact")[0].find_elements_by_tag_name("dd")[3].get_attribute("innerText")
            return (tc_sonu, ad_soyad, durum)

        except:
            return (False, False, False)
    
    
    
    def eDevletGirisiYap(self, tcKimlik="", sifre="", gizliPencere = False):
        
        if tcKimlik == "" and sifre == "":
            tcKimlik = self.kullaniciAdi.text()
            sifre = self.sifre.text()
        
        if tcKimlik == "" or sifre == "":
            return False
        

        url = "https://giris.turkiye.gov.tr/Giris/"
        
        #*E-devlete grup ekleneceği zaman gizliPencere=False olmalı, excel doyası yüklenmiyor yoksa.
        if gizliPencere:
            options = Options()
            options.add_argument('--headless')
            self.br = Firefox(executable_path="geckodriver", options=options)
        else:    
            self.br = Firefox(executable_path="geckodriver")
            #self.br.minimize_window()
            
        
        br= self.br
        br.get(url)
        
        br.find_element_by_id("tridField").send_keys(tcKimlik)
        br.find_element_by_id("egpField").send_keys(sifre)
        br.find_elements_by_class_name("submitButton")[0].click()
        
        self.elemanYuklenmesiniBekle(sinif="userMenuButton")
        
        if len( br.find_elements_by_class_name("userMenuButton") ) > 0:
            return True
        else:
            return False
    
    
    
    def eDevletCikisYap(self):
        self.br.find_elements_by_class_name("userMenuButton")[0].click()
        self.br.find_elements_by_class_name("logout")[0].click()
        self.br.close()
    
    
    
    def eDevlettenSorgula(self, hesKodu, hesKoduKayitli = True):
        

        if self.eDevletGirisiYap() :
            self.pb.setLabelText("E-Devlet girişi yapıldı...")
        else:
            self.yukleniyorAnimasyonuKapat()
            self.mesaj = "E-Devlet girişi başarısız!"
            self.hataMesajiButton.click()
            return
            
            
        br = self.br
        """
        #*TEST EKLEME
        hes1 = "F1D9954714"
        hes2 = "A8U2525219"
        hes3 = "U2A9438817"
        #Alttaki geçersiz hes
        hes4 = "U2U9438817"
        hes5 = "U2T9438817"
        """
        url1 = "https://www.turkiye.gov.tr/saglik-bakanligi-hes-kodu-sorgulama"
        sayac = 0
        br.get(url1)
        
        try:

            br.find_element_by_id("hes_kodu").send_keys(hesKodu)
            br.find_elements_by_class_name("submitButton")[0].click()
                
            
            while(True):
                tc_sonu, ad_soyad, durum = self.sonucGeldimi()
                #Hes kodu geçersiz uyarısı kontrol ediliyor.
                if len(br.find_elements_by_class_name("warningContainer")) > 0 :
                    tc_sonu, ad_soyad, durum = (False, False, False)
                    break
                
                if tc_sonu or sayac > TIMEOUT:
                    break
                sayac += 1
                time.sleep(1)
            
            
            
            #* Burada edevletten sonuç gelmiş ve tc_sonu, hes ve durum değişkenlerine aktarılmış oluyor. 
            #*Sonuç gelmemişse bu değişkenlere False değeri variliyor.
            
            
            self.eDevletCikisYap()
            self.yukleniyorAnimasyonuKapat()
            
            
            if tc_sonu:
                if not hesKoduKayitli:
                    son_uc = tc_sonu[-3:]
                    ogrenciListe = vt.DB.tc_sonu_ile(son_uc)
                    
                    #Öğrenci veritabanında yoksa.
                    if len(ogrenciListe) == 0:
                        self.durum.setText(durum)
                        self.tcNo.setText(tc_sonu)
                        
                        soyad = ad_soyad.split(" ")[-1]
                        self.soyad.setText(soyad)
                        self.ad.setText(ad_soyad.replace(soyad, "").strip())
                        
                    elif len(ogrenciListe) == 1:
                        self.ogrenciGoster(ogrenciListe[0])
                        self.mesaj = "Olası eşleşme"
                        self.hataMesajiButton.click()
                    else:
                        self.ogrenciGoster(ogrenciListe[0])
                        self.tekHesSorguSonucu = ogrenciListe
                        self.mesaj = "Olası eşleşmeler"
                        self.hataMesajiButton.click()
                        
                        #TODO RESMİN ALTINA SIRADAKİ GİBİ BİR BUTON KOY. tek hes sorgu sonucunu bir ileriye alsın.
                
                else:
                    #Hes kodu kayıtlı, öğrenci zaten gösterildi. Durum yazılacak.
                    self.durum.setText(durum)

       
            else:
                self.mesaj = "Hatalı hes kodu!"
                self.hataMesajiButton.click()

             
        except:
            br.close()
            self.yukleniyorAnimasyonuKapat()
            self.mesaj = "Beklenmeyen bir hata oluştu!"
            self.hataMesajiButton.click()
            
          
    #True döndürürse grup kaydedilmiş ve hes kodlarını kaydedilmiştir.
    def grupOlustur(self, grupAdi, tamDosyaYolu, ogrenciler, aciklama="" , varolanGrubaEkle= False):
        

        if not varolanGrubaEkle:
            if not self.eDevletGirisiYap():
                self.mesaj = "E-Devlet girişi yapılamadı."
                self.hataMesajiButton.click()
                return False
            
            br = self.br
            url = "https://www.turkiye.gov.tr/saglik-bakanligi-toplu-hes-kodu-sorgulama?yeni=grup"     
            br.get(url)
            br.find_element_by_id("grupAdi").clear()
            br.find_element_by_id("grupAdi").send_keys(grupAdi)
            br.find_element_by_id("grupAciklama").send_keys(aciklama)
            br.find_elements_by_class_name("submitButton")[0].click()
            
            if not self.elemanYuklenmesiniBekle(id="hesKodu"):
                return False
        
        else:
            self.eDevlettenGrupGuncelleSayfasiAc("öğrenciler", browser=True)
            self.elemanYuklenmesiniBekle(linkText="Hes Kodlarını Excelden Yükle") 
            br = self.br 
              
              
        br.find_element_by_link_text("Hes Kodlarını Excelden Yükle").click()
        self.elemanYuklenmesiniBekle(sinif="filedropzone") 
        
        
        br.find_elements_by_class_name("filedropzone")[0].click()
        subprocess.call(["cscript.exe", "deneme.vbs", tamDosyaYolu ]) 
        
        time.sleep(4)
       
        br.find_element_by_name("btn").click()
        
        
        if self.elemanYuklenmesiniBekle(sinif="list-right"):
            if len( br.find_elements_by_class_name("confirmContainer")) > 0 :
                
                vt.DB.toplu_kayit_ekle(ogrenciler)
                self.mesaj = "Liste kaydedildi."
                self.hataMesajiButton.click()
                br.close()
                self.btnSinifIslemleri.click()
                return True
            else:
                self.mesaj = "Hata: Liste e-devlete kaydedilemedi!"
                self.hataMesajiButton.click()
                return False
        else:
            return False
        
    
    
  
    def eDevlettenGrupGuncelleSayfasiAc(self, grupAdi, browser= True, ogrenci=[], ogrenciGuncellenecek=False):
        if browser:
            if not self.eDevletGirisiYap():
                return

        br = self.br
        url = "https://www.turkiye.gov.tr/saglik-bakanligi-toplu-hes-kodu-sorgulama"
        br.get(url)
        self.elemanYuklenmesiniBekle(id="resultTable")
        
        
        listeler= br.find_elements_by_link_text("Güncelle")
        if len( listeler ) == 0:
            return
        elif len(listeler) == 1:
            br.find_element_by_link_text("Güncelle").click()
        else:
            tumSutunlar = br.find_elements_by_tag_name("td")
            gruplar = [i.text for i in tumSutunlar if tumSutunlar.index(i) % 4 == 0]
            grupIndex = gruplar.index(grupAdi)
            br.find_elements_by_link_text("Güncelle")[grupIndex].click()
        
        #*İlgili grup güncellenmeye hazır.
        self.elemanYuklenmesiniBekle(sinif="actionButton")
        br.find_elements_by_class_name("actionButton")[1].click()
        
        self.elemanYuklenmesiniBekle(sinif="actionButton")

        if len(ogrenci) > 0:
            if self.grubaHesEkle(ogrenci[HES_KODU_INDEX], browserAc=False):
                
                if ogrenciGuncellenecek:
                    vt.DB.veriguncelle(ogrenci[OKUL_NO_INDEX], ogrenci[SINIF_INDEX], ogrenci[HES_KODU_INDEX])
                else:
                    vt.DB.toplu_kayit_ekle([ogrenci])
                    
                    
                self.mesaj = "Öğrenci kaydedildi."
                self.hataMesajiButton.click()
                
    
    
    
    def eDevletHesSil(self, hes, browserAc = True):
        
        if browserAc: 
            self.eDevlettenGrupGuncelleSayfasiAc("öğrenciler")
            
        br = self.br
        
        #*İlk satır tablo başlıkları
        satirlar = br.find_elements_by_tag_name("tr")
        if len(satirlar) > 1:
            satir = satirlar[1]
        
            #*Satırların class hidden durumları değiştiriliyor.Yoksa 1. sayfadan sonrakiler görünmüyor.  
            br.execute_script("arguments[0].setAttribute('class', '')", satir)
        
            sutunListesi = br.find_elements_by_tag_name("td")
            tumHesler = [i.text for i in sutunListesi if sutunListesi.index(i) % 2 == 0]

            br.find_elements_by_link_text("Sil")[tumHesler.index(hes)].click()
            self.elemanYuklenmesiniBekle(sinif="actionButton")



  
    def edevlettenHepsiniSil(self):
        if not self.eDevletGirisiYap():
            return

        br = self.br
        url = "https://www.turkiye.gov.tr/saglik-bakanligi-toplu-hes-kodu-sorgulama"
        br.get(url)
        self.elemanYuklenmesiniBekle(id="resultTable")
        
        while(len(br.find_elements_by_link_text("Sil")) > 0 ):
            
            br.find_element_by_link_text("Sil").click()
            self.elemanYuklenmesiniBekle(sinif="radioButton")
            
            br.find_element_by_class_name("radioButton").click()
            br.find_element_by_class_name("actionButton").click()
            
            self.elemanYuklenmesiniBekle(sinif="contentToolbar")
            br.find_element_by_partial_link_text("Grup Listesi").click()
            
            self.elemanYuklenmesiniBekle(id="resultTable")

        

    def elemanYuklenmesiniBekle(self, id="", sinif="", linkText=""):
        br = self.br
        timeout = TIMEOUT
        
        if id != "":
            
            while len(br.find_elements_by_id(id)) == 0 and timeout > 0:
                time.sleep(1)
                timeout -= 1
            if timeout == 0:
                return False
            else:
                return True
        
        elif sinif != "":
            while len(br.find_elements_by_class_name(sinif)) == 0 and timeout > 0:
                time.sleep(1)
                timeout -= 1
            if timeout == 0:
                return False
            else:
                return True
        
        elif linkText != "":
            while len(br.find_elements_by_link_text(linkText)) == 0 and timeout > 0:
                time.sleep(1)
                timeout -= 1
            if timeout == 0:
                return False
            else:
                return True
    
    

    def sinifIslemleri(self):
        self.comboboxSinifSec.clear()
        siniflar = []
        tmp_siniflar = vt.DB.sinif_isimleri()

        #[(sınıf,),(sınıf,),(sınıf,)] şeklinde geliyor veritabanından.
        for i in tmp_siniflar:
            siniflar.append(*i)
            
        #*Veritabanından çekilen sınıflar comboboxa aktarılıyor.
        self.comboboxSinifSec.addItems(siniflar)
        
        tablo = self.sorguSonucuTableWidget
        #*Resim, ad soyad, tc ve açıklama olacak şekilde 4 sütun.
        tablo.setColumnCount(4)
        tablo.setRowCount(1)
              
        item1 = QTableWidgetItem("Resim")
        item1.setTextAlignment(Qt.AlignCenter)
        item1.setBackground(QColor("#616161"))
        item1.setForeground(QColor("#ffffff"))
        tablo.setItem(0,0,item1) 
        item2 = QTableWidgetItem("Öğrenci")
        item2.setTextAlignment(Qt.AlignCenter)
        item2.setBackground(QColor("#616161"))
        item2.setForeground(QColor("#ffffff"))
        tablo.setItem(0,1,item2)        
        item3 = QTableWidgetItem("Okul No")
        item3.setTextAlignment(Qt.AlignCenter)
        item3.setBackground(QColor("#616161"))
        item3.setForeground(QColor("#ffffff"))
        tablo.setItem(0,2,item3)     
        item4 = QTableWidgetItem("Açıklama")
        item4.setTextAlignment(Qt.AlignCenter)
        item4.setBackground(QColor("#616161"))
        item4.setForeground(QColor("#ffffff"))
        tablo.setItem(0,3,item4)
        tablo.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tablo.horizontalHeader().hide()
        tablo.verticalHeader().hide()
      
        
    
    def btnSinifSorgulaClicked(self):
        
        if self.kullaniciAdi.text() != "" and self.sifre.text() != "":
            self.sorguSonucuLabel.setText("")
            self.yukleniyorAnimasyonuBaslat()
            th = Thread(target= self.sinifSorgula)
            th.start()

        else:
            self.mesaj = "E-Devlet kullanıcı adı ve şifrenizi girmelisiniz!"
            self.hataMesajiButton.click()

        
        
     
    
    def sinifSorgula(self):
        
        self.tabloyaGonderilenOgrenciler = []
        self.tabloyaGonderilenSorunlar = []
        
        sinif = self.comboboxSinifSec.currentText()
        
        if not self.eDevletGirisiYap():
            self.yukleniyorAnimasyonuKapat()
            self.br.close()
            self.mesaj = "E-Devlet girişi yapılamadı!"
            self.hataMesajiButton.click()
            return
        
        

        br = self.br
        br.get("https://www.turkiye.gov.tr/saglik-bakanligi-toplu-hes-kodu-sorgulama") 
        self.elemanYuklenmesiniBekle(linkText="Yeni Grup Oluştur")     
        
        url = "https://www.turkiye.gov.tr/saglik-bakanligi-toplu-hes-kodu-sorgulama?i=s&s=0"
        br.get(url)
        self.elemanYuklenmesiniBekle(sinif="compact")
        

        gecersizler = br.find_elements_by_css_selector("table[summary='Geçersiz HES Kodu Listesi (Süresi dolmuş, Silinmiş)']")
        pozitifler = br.find_elements_by_css_selector("table[summary='Riskli HES Kodu Listesi (Pozitif, Temaslı)']")

        if len(gecersizler) != 0 or len(pozitifler) != 0:  
           #Tüm satırlar görünür hale getiriliyor.
            for eleman in br.find_elements_by_tag_name("tr"):
                br.execute_script("arguments[0].setAttribute('class','')", eleman)
        
        
            ogrenciler = []
            hesler = []
            sorunlar = []
            if len(pozitifler) > 0:
                self.sorguSonucuLabel.setText("Sorun var!")
                #hes_kodu, tc_no, açıklama
                tumBilgiler =[ i.text for i in pozitifler[0].find_elements_by_tag_name("td") ] 
                #*Bunlar edevletten gelecek sonuçlar.
                tmp_hesler = [i for i in tumBilgiler if tumBilgiler.index(i) % 3 == 0]
                tmp_sorunlar = [i for i in tumBilgiler if tumBilgiler.index(i) % 3 == 2]
                #*Sınıf sorgulandığı için hesler zaten veritabanında kayıtlı.
                #Bu sorguda sorun çıkarsa döngü ile veritabanından sorgulanabilir.
                
                
                tmp_ogrenciler = vt.DB.hes_kodu_listesi_ile(hesler)
                if len(tmp_ogrenciler) > 0:
                    for ogrenci in tmp_ogrenciler:
                        if ogrenci[SINIF_INDEX] == sinif:
                            ind = tmp_ogrenciler.index(ogrenci)
                            ogrenciler.append(ogrenci)
                            hesler.append(tmp_hesler[ind])
                            sorunlar.append(tmp_sorunlar[ind])

                
    
            if len(gecersizler) > 0:
                self.sorguSonucuLabel.setText("Sorun var!")
                tumBilgiler =[ i.text for i in gecersizler[0].find_elements_by_tag_name("td")  ]

                
                tmp_hesler = [i for i in tumBilgiler if tumBilgiler.index(i) % 3 == 0]
                tmp_sorunlar = [i for i in tumBilgiler if tumBilgiler.index(i) % 3 == 1]
                #*Sınıf sorgulandığı için hesler zaten veritabanında kayıtlı.
                #Bu sorguda sorun çıkarsa döngü ile veritabanından sorgulanabilir.
                tmpOgrenciler = vt.DB.hes_kodu_listesi_ile(tmp_hesler)
                    
                if len(tmpOgrenciler) > 0:
                    for ogrenci in tmpOgrenciler:
                        if ogrenci[SINIF_INDEX] == sinif:
                            ind = tmpOgrenciler.index(ogrenci)
                            ogrenciler.append(ogrenci)
                            hesler.append(tmp_hesler[ind])
                            sorunlar.append(tmp_sorunlar[ind])
                
              
            
            self.tabloyaGonderilenOgrenciler = ogrenciler
            self.tabloyaGonderilenSorunlar = sorunlar
            
        else:
            self.sorguSonucuLabel.setText("Sorun yok.")
        
        
        
        self.tabloGonderButton.click()          
        self.animasyonKapatButon.click()   
        br.close()
    
    
    def tabloIsle(self) :
        
        if len(self.tabloyaGonderilenOgrenciler) > 0:
            self.sorguSonucuLabel.setStyleSheet("QLabel{background-color:red}")
    
            ogrenciler = self.tabloyaGonderilenOgrenciler
            sorunlar = self.tabloyaGonderilenSorunlar
            
            tablo = self.sorguSonucuTableWidget
            tablo.setRowCount(len(ogrenciler)+1)

            for ogrenci in ogrenciler:
                sira = ogrenciler.index(ogrenci) + 1
                fotoLabel = QLabel()
                resim = ogrenci[RESIM_INDEX].scaled(45,50, Qt.KeepAspectRatio)
                fotoLabel.setPixmap(resim)


                tablo.setCellWidget(sira, 0, fotoLabel) 
                tablo.setItem(sira, 1, QTableWidgetItem(ogrenci[AD_INDEX] + " " + ogrenci[SOYAD_INDEX]))
                tablo.setItem(sira, 2, QTableWidgetItem(str(ogrenci[OKUL_NO_INDEX])))
                tablo.setItem(sira, 3, QTableWidgetItem(sorunlar[sira-1]))
                    
                    
                #* Tablo biçimlendirmeleri
                tablo.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
                tablo.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
                tablo.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
                tablo.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
                tablo.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        else:
            self.sorguSonucuLabel.setStyleSheet("QLabel{background-color:green}")
            self.sorguSonucuLabel.setText("Sorun yok")
    
    def dosyaSec(self) :
        dosyaAdi, _ = QFileDialog.getOpenFileName(self, 'Dosya seç', '', "Excel files (*.xls *.xlsx)")
        if dosyaAdi :
            self.dosyaYoluLineEdit.setText(dosyaAdi)
        
        
        
    def dizinSec(self):
        fotoDizin = str(QFileDialog.getExistingDirectory(self, "Dizin seç"))
        if fotoDizin:
            self.fotoYoluLineEdit.setText(fotoDizin)
    
    
   
    def topluKaydet(self):
        
        if self.kullaniciAdi.text() == "" or self.sifre.text() == "":
            self.mesaj = "E-Devlet kullanıcı adı ve şifrenizi girelisiniz."
            self.hataMesajiButton.click()
            return
        
        if os.path.exists(self.dosyaYoluLineEdit.text()) and os.path.exists(self.fotoYoluLineEdit.text()):
            #Excelden veriler alınıyor.
            ogrenciler = self.exceldenVeriAl()
            #Veritabanına kaydediliyor.
            
            hesListesi = [ogrenci[HES_KODU_INDEX] for ogrenci in ogrenciler]
            
            self.hesExcelDosyalasiniKaydet(hesListesi)
            
            #Tek grup oluşturuldu. Başarılı olursa veritabanına kaydedilecek.
            #Grup yoksa oluşturulacak varsa üstüne hes kodları eklenecek.
            yol = os.getcwd() + os.sep + "ogrenciler.xlsx"
            if len(vt.DB.sinif_isimleri()) == 0:
                th = Thread(target=self.grupOlustur, args=("öğrenciler", yol, ogrenciler))
                th.start()
            else:
                th = Thread(target=self.grupOlustur, args=("öğrenciler", yol, ogrenciler, "", True))
                th.start()
            
            
        else:
            self.mesaj = "Excel dosyasını ve fotoğraf dizinini seçmelisiniz!"
            self.hataMesajiButton.click()
    
    
    
    
    def exceldenVeriAl(self):
        dosyaYolu = self.dosyaYoluLineEdit.text()
        #TODO Şablon excel dosyasına göre indexler düzenlenecek.
        indexler = {"ad":1, "soyad":2, "tc_no":3, "hes":5, "okul_no":0, "sinif":4}
        
        wb = load_workbook(dosyaYolu, read_only=True)
        ws = wb.active

        fotoYol =  self.fotoYoluLineEdit.text()
        dosyaIsimleri = os.listdir(fotoYol)
        dosyaTamIsimleri = [i for i in dosyaIsimleri if i.endswith(".jpg") or i.endswith(".png")]

        veriler = ws.values
        
        ogrenciler = []
        sayac = -1
        for satir in veriler:
            #Excelde ilk satır başlık satırı.
            sayac += 1
            if sayac == 0:
                continue
            
            ogrenci = []
            tmp = [*satir]
            
            tc = tmp[ indexler["tc_no"] ]
            if tc == None:
                break
            ogrenci.append(str(tc))
            
            ad = tmp[ indexler["ad"] ]
            ogrenci.append(ad)
            
            soyad = tmp[ indexler["soyad"] ]
            ogrenci.append(soyad)
            
            numara =  tmp[ indexler["okul_no"] ]
            ogrenci.append( numara )
            
            numaraString = str(numara)
            
            sinif = tmp[ indexler["sinif"] ]
            ogrenci.append(sinif)
            
            if numaraString+".jpg" in dosyaTamIsimleri:
                resim = fotoYol + "/" + numaraString + ".jpg"
            elif numaraString+".png" in dosyaIsimleri:
                resim = fotoYol + "/" + numaraString + ".png"
            else:
                resim = ""  
                
            ogrenci.append(resim)
            hes = tmp[ indexler["hes"] ]
            ogrenci.append(hes)
            ogrenci.append(1)  
            ogrenciler.append(ogrenci)
            
        wb.close()
        
        return ogrenciler 
    
    
    
    
    def hesExcelDosyalasiniKaydet(self, hesListesi):
    
        try:
            wb = xlsxwriter.Workbook('ogrenciler.xlsx')
            ws = wb.add_worksheet()
            ws.write(0,0,"HES KODU")
            
            yazilacak_satir = 1
            for hes in hesListesi:
                ws.write(yazilacak_satir, 0, hes )
                yazilacak_satir += 1
                    
            wb.close()
            return True
        
        except:
            return False
        
    
        
    
    def sinifSilClicked(self):
        
        qm = QMessageBox(self)
        mesaj = f"{self.comboboxSinifSec.currentText()} silinecek. Onaylıyor musunuz?"
        ret = qm.warning(self,'Öğrenci Sil', mesaj, qm.Yes | qm.No)

        if ret == qm.Yes:
            vt.DB.sinif_sil( self.comboboxSinifSec.currentText() )
            self.btnSinifIslemleri.click()
            self.mesaj = "Sınıf silindi."
            self.hataMesajiButton.click()
    
        

if __name__  == "__main__":
    
    app = QApplication(sys.argv)
    window = MWindow()
    sys.exit( app.exec_() )